import datetime
import os
import traceback
from io import BytesIO
import openpyxl
import requests
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
from datetime import timedelta
import chime






class Prep_records():
    def make_heat_map(self):
        mas = np.array([[0] * 1320] * 760)
        response = requests.get(f'{self.res.get("faces")[0].get("photo")}')
        datafile = Image.open(BytesIO(response.content))
        datafile.save(f'{self.dir}/orig_photo/{self.camera_name}.png')

        if len(self.res.get("faces")) > self.count_results:
            self.progressBar.setMaximum(self.count_results)
        else:
            self.progressBar.setMaximum(len(self.res.get("faces")))

        for count, face in enumerate(self.res.get("faces")):
            self.progressBar.setValue(count)
            if not self.Flag:
                return
            zone = [face.get("bbox").get("left"), face.get("bbox").get("top"),
                    face.get("bbox").get("right"), face.get("bbox").get("bottom")]
            zone = [(int(zone[0]) + 20), (int(zone[1]) + 20), (int(zone[2]) + 20), (int(zone[3]) + 20)]
            centr = int((zone[2] - zone[0]) / 2)
            a = 10.0
            for j in range(zone[0] + centr, zone[0], -1):
                b = a
                for i in range(zone[1] + centr, zone[1], -1):
                    mas[i, j] += a
                    a = a / 1.01
                a = b / 1.01

            a = 10.0 / 1.01
            for j in range(zone[0] + centr + 1, zone[2]):
                b = a
                for i in range(zone[1] + centr - 1, zone[1], -1):
                    mas[i, j] += a
                    a = a / 1.01
                a = b / 1.01

            a = 10.0 / 1.01
            for j in range(zone[0] + centr - 1, zone[0], -1):
                b = a
                for i in range(zone[1] + centr + 1, zone[3]):
                    mas[i, j] += a
                    a = a / 1.01
                a = b / 1.01

            a = 10.0
            for j in range(zone[0] + centr, zone[2]):
                b = a
                for i in range(zone[1] + centr, zone[3]):
                    mas[i, j] += a
                    a = a / 1.01
                a = b / 1.01
            mas[zone[1] + centr, zone[0] + centr] -= 1
            if count == self.count_results:
                break

        plt.imshow(mas, cmap='inferno', interpolation='nearest')
        plt.axis('off')
        plt.savefig(f'{self.dir}/heat_map/{self.camera_name}.png', dpi=300, transparent=True)
        self.creating_imeges()


    def make_report(self):
        self.quality = 0
        self.count_faces = 0
        self.bbox = 0
        for count, face in enumerate(self.res.get("faces")):
            try:
                self.bbox += (face.get("bbox").get("right") - face.get("bbox").get("left")) * (
                        face.get("bbox").get("bottom") - face.get("bbox").get("top"))
                if face.get("features").get("quality").get("synesis", 0) > 0:
                    self.quality += face.get("features").get("quality").get("synesis")
                    self.count_faces += 1
                if face.get("features").get("quality").get("visionlabs", 0) > 0:
                    self.quality += face.get("features").get("quality").get("visionlabs")
                    self.count_faces += 1
                if face.get("features").get("quality").get("tevian", 0) > 0:
                    self.quality += face.get("features").get("quality").get("tevian")
                    self.count_faces += 1
                if face.get("features").get("quality").get("ntechlab", 0) > 0:
                    self.quality += face.get("features").get("quality").get("ntechlab")
                    self.count_faces += 1
            except:
                pass
        self.count = count


    def save_report_only(self):
        Wordbook = openpyxl.Workbook()
        sheet = Wordbook['Sheet']
        sheet.title = "Отчет"
        for column, column_name in enumerate(["id камеры","Имя камеры","Количество детектов","Средний размер лица","Среднее качество"]):
            sheet.cell(1, column+1).value = column_name
        time_start = (datetime.datetime.now() - timedelta(self.days, hours=3)).strftime("%Y-%m-%dT%H:%M:%SZ")
        self.show_bar()
        self.progressBar.setMaximum(len(self.cameras))
        for index, cam_id in enumerate(self.cameras):
            if not self.Flag:
                return
            try:
                sheet.cell(sheet.max_row + 1, 1).value = cam_id
                self.camera_name = self.get_cam_name(cam_id)
                if not self.camera_name:
                    sheet.cell(sheet.max_row, 2).value = "Не нашел такую камеру"
                    continue
                self.res = self.get_quality_detects(cam_id, 10000, time_start)
                self.progressBar.setValue(index)
                if self.res.get("faces") == []:
                    sheet.cell(sheet.max_row, 2).value = self.camera_name
                    sheet.cell(sheet.max_row, 3).value = "Нет детектов"
                    continue
                self.make_report()

                sheet.cell(sheet.max_row, 2).value = self.camera_name
                sheet.cell(sheet.max_row, 3).value = len(self.res.get("faces"))
                sheet.cell(sheet.max_row, 4).value = int(self.bbox / self.count)
                sheet.cell(sheet.max_row, 5).value = self.quality/self.count_faces



            except Exception as ex:
                traceback.print_exc()
                sheet.cell(sheet.max_row, 6).value = "Ошибка обработки"
            finally:
                self.label_8.setText(f"Готово {index+1} / {len(self.cameras)}")

        Wordbook.save(f'{self.dir}/Качество лиц_{datetime.datetime.now().strftime("%d.%m.%Y %H.%M.%S")}.xlsx')
        self.hide_bar()
        self.textBrowser.append("Готово!")
        chime.success()


    def save_map_only(self):
        self.dir = f"{self.dir}/report_{datetime.datetime.now().strftime('%d.%m.%Y %H.%M.%S')}"
        os.makedirs(f"{self.dir}/heat_map")
        os.mkdir(f"{self.dir}/orig_photo")
        os.mkdir(f"{self.dir}/map_with_backing")
        time_start = (datetime.datetime.now() - timedelta(self.days, hours=3)).strftime("%Y-%m-%dT%H:%M:%SZ")
        self.show_bar()
        for index, cam_id in enumerate(self.cameras):
            if not self.Flag:
                return
            try:
                self.res = self.get_quality_detects(cam_id, 10000, time_start)
                self.camera_name = self.get_cam_name(cam_id)
                if not self.camera_name:
                    self.textBrowser.append(f"{cam_id} - Не нашел такую камеру")
                    continue

                if self.res.get("faces") == []:
                    self.textBrowser.append(f"{cam_id} - Нет детектов")
                    continue

                self.make_heat_map()
            except:
                self.textBrowser.append(f"{cam_id} - Ошибка обработки")
            finally:
                self.label_8.setText(f"Готово {index+1} / {len(self.cameras)}")

        self.hide_bar()
        self.textBrowser.append("Готово!")
        chime.success()


    def save_chart(self):
        self.dir = f"{self.dir}/report_{datetime.datetime.now().strftime('%d.%m.%Y %H.%M.%S')}"
        os.makedirs(f"{self.dir}/heat_map")
        os.mkdir(f"{self.dir}/orig_photo")
        os.mkdir(f"{self.dir}/map_with_backing")
        Wordbook = openpyxl.Workbook()
        sheet = Wordbook['Sheet']
        sheet.title = "Отчет"
        sheet.cell(1, 1).value = "id камеры"
        sheet.cell(1, 2).value = "Имя камеры"
        sheet.cell(1, 3).value = "Количество детектов"
        sheet.cell(1, 4).value = "Средний размер лица"
        sheet.cell(1, 5).value = "Среднее качество"
        time_start = (datetime.datetime.now() - timedelta(self.days, hours=3)).strftime("%Y-%m-%dT%H:%M:%SZ")
        self.show_bar()

        for index, cam_id in enumerate(self.cameras):
            if not self.Flag:
                return
            try:
                sheet.cell(sheet.max_row + 1, 1).value = cam_id
                self.camera_name = self.get_cam_name(cam_id)
                if not self.camera_name:
                    sheet.cell(sheet.max_row, 2).value = "Не нашел такую камеру"
                    continue

                self.res = self.get_quality_detects(cam_id, 10000, time_start)

                if self.res.get("faces") == []:
                    sheet.cell(sheet.max_row, 2).value = self.camera_name
                    sheet.cell(sheet.max_row, 3).value = "Нет детектов"
                    continue

                self.make_heat_map()
                self.make_report()

                sheet.cell(sheet.max_row, 2).value = self.camera_name
                sheet.cell(sheet.max_row, 3).value = len(self.res.get("faces"))
                sheet.cell(sheet.max_row, 4).value = int(self.bbox / (self.count + 1))
                sheet.cell(sheet.max_row, 5).value = self.quality / self.count_faces
                sheet.cell(sheet.max_row, 6).value = '=HYPERLINK("{0}", "{1}")'.format(f"orig_photo/{self.camera_name}.png", "фото с камеры")
                sheet.cell(sheet.max_row, 7).value = '=HYPERLINK("{0}", "{1}")'.format(f"heat_map/{self.camera_name}.png", "тепловая карта")
                sheet.cell(sheet.max_row, 8).value = '=HYPERLINK("{0}", "{1}")'.format(f"map_with_backing/{self.camera_name}.png", "карта с подложкой")


            except Exception as ex:
                traceback.print_exc()
                sheet.cell(sheet.max_row, 2).value = "Ошибка обработки"
            finally:
                self.label_8.setText(f"Готово {index+1} / {len(self.cameras)}")

        Wordbook.save(f'{self.dir}/Качество лиц.xlsx')
        self.hide_bar()
        self.textBrowser.append("Готово!")
        chime.success()


    def creating_imeges(self):
        heat = Image.open(f'{self.dir}/heat_map/{self.camera_name}.png')
        heat_crop = heat.crop((260, 319, 1707, 1135))
        width = 1280
        height = 720
        resized_heat = heat_crop.resize((width, height), Image.LANCZOS)
        orig = Image.open(f'{self.dir}\\orig_photo\\{self.camera_name}.png')
        heat_load = resized_heat.load()
        orig_load = orig.load()
        for j in range(1280):
            for k in range(720):
                orig_load[j, k] = heat_load[j, k]
        orig.save(f'{self.dir}\\heat_map\\{self.camera_name}.png')
        heat = Image.open(f'{self.dir}\\heat_map\\{self.camera_name}.png')
        orig = Image.open(f'{self.dir}\\orig_photo\\{self.camera_name}.png')
        new_img = Image.blend(heat, orig, 0.5)
        new_img.save(f'{self.dir}\\map_with_backing\\{self.camera_name}.png')